import ecommerce.shipping
# ecommerce.shipping.calculate_shipping()

import random

# for i in range(3):
    # print(random.randint(10,20))

members = ["John","marry","max","Mosh"]
leader = random.choice(members)
print(leader)


class Dice:
    def roll(self):
        first = random.randint(1,6)
        second  = random.randint(1,6)
        return first,second

dice = Dice()
print(dice.roll())

from pathlib import Path

path = Path()
# print(path.glob('*.py'))
for file in path.glob('*.py'):
    print(file)

    import openpyxl as xl
    wb = xl.load_workbook("transactions.xlsx")
    Sheet = wb["Sheet1"]
    cell = Sheet["a1"]
    print(cell.value)

for row in range(1,Sheet.max_row+1):
    cell = Sheet.cell(row,1)
    print(cell.value)

# vowels list
vowels = [2,6,1,4,8,9,0]

# sort the vowels
vowels.sort(reverse=False)

# print vowels
print('Sorted list (in Descending):', vowels)

data_list = [-5, -23, 5, 0, 23, -6, 23, 67]
new_list = []

while data_list:
    minimum = data_list[0]  # arbitrary number in list 
    for x in data_list: 
        if x < minimum:
            minimum = x
    new_list.append(minimum)
    data_list.remove(minimum)    

print (new_list)


numberArray = [5,2,3,6,7,1,2,9,0]
newlist= []
while numberArray:
    min= numberArray[0]
    for x in numberArray:
        if x>min:
            min = x
    newlist.append(min)
    numberArray.remove(min)
print(newlist)

